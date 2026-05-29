from decimal import Decimal
from django.db.models import Sum, Max, Q
from django.db.models.functions import Coalesce

from finances.models import Invoice, InvoicePayment, OutgoingPayment, PurchaseOrder
from django.db.models import F
from Project.models import ProjectBudget
def get_financial_summary(filters):
    budget = ProjectBudget.objects.aggregate(
        total=Coalesce(Sum("total_budget"), Decimal("0.00"))
    )["total"]

    invoiced = Invoice.objects.exclude(
        status="Cancelled"
    ).aggregate(
        total=Coalesce(Sum("total_amount"), Decimal("0.00"))
    )["total"]

    received = InvoicePayment.objects.aggregate(
        total=Coalesce(Sum("amount"), Decimal("0.00"))
    )["total"]

    expenses = OutgoingPayment.objects.aggregate(
        total=Coalesce(Sum("amount"), Decimal("0.00"))
    )["total"]

    return {
        "revenue": invoiced,
        "expenses": expenses,
        "profit": received - expenses,
    }

def get_all_tab_data(filters):
    """
    Data for 'All' tab (Dashboard overview)
    Shows summary cards + recent highlights
    """
    from django.db.models import Max
    
    qs = Invoice.objects.select_related("client", "project").exclude(status="Cancelled")

    if filters.get("from_date"):
        qs = qs.filter(issue_date__gte=filters["from_date"])

    if filters.get("to_date"):
        qs = qs.filter(issue_date__lte=filters["to_date"])

    if filters.get("client"):
        qs = qs.filter(client_id=filters["client"])

    if filters.get("project"):
        qs = qs.filter(project_id=filters["project"])

    # Group by client and project using a dict
    grouped = {}
    for invoice in qs:
        key = (invoice.client_id, invoice.project_id)
        if key not in grouped:
            grouped[key] = {
                "client": invoice.client,
                "project": invoice.project,
                "invoices": []
            }
        grouped[key]["invoices"].append(invoice)
    
    all_rows = []
    total_revenue = Decimal("0.00")
    total_expenses = Decimal("0.00")
    total_profit = Decimal("0.00")
    total_outstanding = Decimal("0.00")
    
    for (client_id, project_id), group_data in grouped.items():
        invoices = group_data["invoices"]
        client = group_data["client"]
        project = group_data["project"]
        
        # Calculate totals for this client/project group
        total_invoiced = sum(inv.total_amount for inv in invoices)
        
        expenses = OutgoingPayment.objects.filter(
            vendor_bill__purchase_order__project=project
        ).aggregate(
            total=Coalesce(Sum("amount"), Decimal("0.00"))
        )["total"] if project else Decimal("0.00")
        
        received = InvoicePayment.objects.filter(
            invoice__client=client,
            invoice__project=project
        ).aggregate(
            total=Coalesce(Sum("amount"), Decimal("0.00"))
        )["total"]
        
        outstanding = total_invoiced - (received or Decimal("0.00"))
        profit = (received or Decimal("0.00")) - (expenses or Decimal("0.00"))
        profit_margin = (profit / total_invoiced * 100) if total_invoiced > 0 else Decimal("0.00")
        
        # Aggregate for summary cards
        total_revenue += total_invoiced
        total_expenses += (expenses or Decimal("0.00"))
        total_profit += profit
        total_outstanding += outstanding
        
        last_payment_date = InvoicePayment.objects.filter(
            invoice__client=client,
            invoice__project=project
        ).aggregate(
            last_date=Max("payment_date")
        )["last_date"]
        
        # Get latest invoice status for this client/project
        latest_invoice = max(invoices, key=lambda x: x.issue_date)
        
        all_rows.append({
            "client_name": client.company_name if client else "N/A",
            "project_name": project.project_name if project else "N/A",
            "total_revenue": float(total_invoiced),
            "total_expenses": float(expenses or Decimal("0.00")),
            "profit": float(profit),
            "profit_margin_percent": float(profit_margin),
            "latest_invoice_status": latest_invoice.status or "Draft",
            "last_receipt_date": last_payment_date,
            "outstanding_amount": float(outstanding),
            "collection_rate_percent": float((received or Decimal("0.00")) / total_invoiced * 100) if total_invoiced > 0 else 0,
        })

    # Get top 5 recent highlights (sorted by last receipt date or latest invoice)
    from datetime import date
    recent_highlights = sorted(
        all_rows, 
        key=lambda x: x["last_receipt_date"] or date.min, 
        reverse=True
    )[:5]

    # Summary cards
    summary_cards = {
        "total_revenue": float(total_revenue),
        "total_expenses": float(total_expenses),
        "total_profit": float(total_profit),
        "total_outstanding": float(total_outstanding),
        "overall_collection_rate": float((total_revenue - total_outstanding) / total_revenue * 100) if total_revenue > 0 else 0,
    }

    return {
        "summary_cards": summary_cards,
        "recent_highlights": recent_highlights,
    }

def get_financial_tab_data(filters):
    """
    Financial Reports tab
    Shows detailed invoice-level records with payment tracking
    """
    qs = Invoice.objects.select_related("client").exclude(status="Cancelled")

    if filters.get("from_date"):
        qs = qs.filter(issue_date__gte=filters["from_date"])

    if filters.get("to_date"):
        qs = qs.filter(issue_date__lte=filters["to_date"])

    if filters.get("client"):
        qs = qs.filter(client_id=filters["client"])

    data = []
    for invoice in qs.order_by("-issue_date"):
        received = InvoicePayment.objects.filter(
            invoice=invoice
        ).aggregate(
            total=Coalesce(Sum("amount"), Decimal("0.00"))
        )["total"]
        
        outstanding = invoice.total_amount - (received or Decimal("0.00"))
        
        last_payment_date = InvoicePayment.objects.filter(
            invoice=invoice
        ).aggregate(
            last_date=Max("payment_date")
        )["last_date"]
        
        # Calculate days overdue if applicable
        from django.utils import timezone
        from datetime import timedelta
        days_overdue = 0
        if invoice.status == "Overdue" and invoice.due_date:
            days_overdue = (timezone.now().date() - invoice.due_date).days
        
        data.append({
            "invoice_no": invoice.invoice_no,
            "invoice_date": invoice.issue_date,
            "client_name": invoice.client.company_name if invoice.client else "N/A",
            "invoice_total": float(invoice.total_amount),
            "amount_paid": float(received or Decimal("0.00")),
            "outstanding_balance": float(outstanding),
            "status": invoice.status,
            "due_date": invoice.due_date,
            "last_payment_date": last_payment_date,
            "days_overdue": days_overdue,
        })

    return {"rows": data}

def get_project_tab_data(filters):
    qs = ProjectBudget.objects.select_related("project")

    if filters.get("project"):
        qs = qs.filter(project_id=filters["project"])

    data = []

    for pb in qs:
        # Skip ProjectBudget records where the related project has been deleted
        if pb.project is None:
            continue

        invoiced = Invoice.objects.filter(
            project=pb.project
        ).aggregate(
            total=Coalesce(Sum("total_amount"), Decimal("0.00"))
        )["total"]

        received = InvoicePayment.objects.filter(
            invoice__project=pb.project
        ).aggregate(
            total=Coalesce(Sum("amount"), Decimal("0.00"))
        )["total"]

        expenses = OutgoingPayment.objects.filter(
            vendor_bill__purchase_order__project=pb.project
        ).aggregate(
            total=Coalesce(Sum("amount"), Decimal("0.00"))
        )["total"]

        data.append({
            "project_no": pb.project.project_no,
            "project_name": pb.project.project_name,
            "project_status": pb.project.status,
            "budget": pb.total_budget or Decimal("0.00"),
            "invoiced": invoiced,
            "received": received,
            "expenses": expenses,
            "profit": received - expenses,
        })

    return {"rows": data}

def get_payment_tab_data(filters):
    qs = InvoicePayment.objects.select_related(
        "invoice", "invoice__client"
    )

    if filters.get("from_date"):
        qs = qs.filter(payment_date__gte=filters["from_date"])

    if filters.get("to_date"):
        qs = qs.filter(payment_date__lte=filters["to_date"])

    return {
        "rows": list(
            qs.values(
                "payment_date",
                "invoice__invoice_no",
                "invoice__client__company_name",
                "payment_method",
                "amount",
                "invoice__status",
            )
        )
    }

def get_po_invoice_tab_data(filters):
    qs = PurchaseOrder.objects.select_related("vendor")

    qs = qs.annotate(
        paid=Coalesce(
            Sum("vendorbill__payments__amount"),
            Decimal("0.00")
        ),
        balance=F("total_amount") - Coalesce(
            Sum("vendorbill__payments__amount"),
            Decimal("0.00")
        )
    )

    return {
        "rows": list(
            qs.values(
                "po_no",
                "vendor__name",
                "total_amount",
                "paid",
                "balance",
            )
        )
    }


from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from django.db.models import Sum

# Cell border style
thin = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def generate_financial_excel(project, invoices, expenses, date_from, date_to):
    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Report"

    # ---------------- TITLE ----------------
    ws.merge_cells("A1:F1")
    ws["A1"] = "FINANCIAL MANAGEMENT SYSTEM"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Project: {project.name}"
    ws["A2"].font = Font(bold=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:F3")
    ws["A3"] = f"Period: {date_from} to {date_to}"
    ws["A3"].alignment = Alignment(horizontal="center")

    # ---------------- CALCULATIONS ----------------
    total_revenue = invoices.filter(status="PAID").aggregate(
        total=Sum("amount")
    )["total"] or 0

    outstanding = invoices.filter(status="PENDING").aggregate(
        total=Sum("amount")
    )["total"] or 0

    total_expenses = expenses.aggregate(
        total=Sum("amount")
    )["total"] or 0

    profit = total_revenue - total_expenses

    summary = [
        ("TOTAL REVENUE", total_revenue),
        ("TOTAL EXPENSES", total_expenses),
        ("PROFIT", profit),
        ("OUTSTANDING", outstanding),
    ]

    # ---------------- SUMMARY TABLE ----------------
    row = 5
    for label, value in summary:
        ws.merge_cells(f"A{row}:C{row}")
        ws.merge_cells(f"D{row}:F{row}")

        ws[f"A{row}"] = label
        ws[f"D{row}"] = value

        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = thin
        ws[f"D{row}"].border = thin
        row += 1

    # ---------------- INVOICES ----------------
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "INVOICES"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1

    ws["A" + str(row)] = "Date"
    ws["B" + str(row)] = "Amount"
    ws["C" + str(row)] = "Status"

    for col in ["A", "B", "C"]:
        ws[f"{col}{row}"].font = Font(bold=True)
        ws[f"{col}{row}"].border = thin

    row += 1
    for inv in invoices:
        ws[f"A{row}"] = inv.created_at.date()
        ws[f"B{row}"] = inv.amount
        ws[f"C{row}"] = inv.status
        for col in ["A", "B", "C"]:
            ws[f"{col}{row}"].border = thin
        row += 1

    # ---------------- EXPENSES ----------------
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "EXPENSES"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1

    ws["A" + str(row)] = "Date"
    ws["B" + str(row)] = "Amount"

    for col in ["A", "B"]:
        ws[f"{col}{row}"].font = Font(bold=True)
        ws[f"{col}{row}"].border = thin

    row += 1
    for exp in expenses:
        ws[f"A{row}"] = exp.created_at.date()
        ws[f"B{row}"] = exp.amount
        for col in ["A", "B"]:
            ws[f"{col}{row}"].border = thin
        row += 1

    return wb